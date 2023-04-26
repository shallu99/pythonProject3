import os
import time

import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By

import openpyxl




class Test_do():
    # driver=None

    # @pytest.fixture()
    # def Test_box(self, browser):
    #     self.driver = browser
    #     # self.driver.implicitly_wait(10)



    print("test completed")

    def read_excel_data(Testdata):
            wb = openpyxl.load_workbook("C:\\Users\\Shreyank M\\OneDrive\\Desktop\\guru99 data.xlsx")
            sheet = wb[Testdata]
            rows = sheet.max_row
            cols = sheet.max_column
            data = []
            for r in range(2, rows + 1):
                row_data = {}
                for c in range(1, cols + 1):
                    row_data[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
                data.append(row_data)
                print(data)
            return data

    @pytest.mark.parametrize("data", read_excel_data("Testdata"))
    def test_login(self, driver, data, ):
        try:

            driver.find_element(By.XPATH, "//input[@name='uid']").send_keys(data["User ID"])

            driver.find_element(By.XPATH, "//input[@name='password']").send_keys(data["Password"])

            driver.find_element(By.XPATH, "//input[@name='btnLogin']").click()
            time.sleep(3)
            assert driver.current_url == "https://demo.guru99.com/V1/html/Managerhomepage.php"
            # name = r"C:\Users\Shreyank M\PycharmProjects\pythonProject2\reports\\attach.png"
            # self.driver.get_screenshot_as_file(name)
            time.sleep(3)
        except AssertionError as error:
            print(error)
            # assert False


        except Exception as error:
            print(error)


            assert False

            # timestamp = str(int(time.time()))
            # directory = "screenshots"
            # if not os.path.exists(directory):
            #     os.makedirs(directory)
            # filename = f"{directory}/screenshot-{timestamp}.png"
            # self.driver.save_screenshot(filename)
            # print(f"Screenshot saved as {filename}")

