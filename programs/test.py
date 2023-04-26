
import os
import time

import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
# from programs.conftest import conftest
import openpyxl
class Test_do():
    driver = None

    @pytest.fixture()
    def Test_box(self):
        # D = conftest()
        # self.driver=D.browser()
        self.driver = webdriver.Chrome("C:\\webdriver\\chromedriver_win32\\chromedriver.exe")
        self.driver.get("https://demo.guru99.com/V1/index.php")
        self.driver.implicitly_wait(10)

        yield

        self.driver.quit()
        print("test completed")

    def read_excel_data(Testdata):
        wb = openpyxl.load_workbook("C:\\Users\\Shreyank M\\OneDrive\\Desktop\\guru99 data.xlsx")
        sheet = wb[Testdata]
        rows = sheet.max_row
        cols = sheet.max_column
        matchingRow=[]
        data = []
        for r in range(2, rows + 1):
            row_data = {}
            for c in range(1, cols + 1):
                row_data[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
            data.append(row_data)
            print(data)
        # Find the rows that contain the search term
        # executes = "Yes"
        # for row in matchingRow:
        #     for cell in row:
        #         if cell.value == executes:
        #             print(cell.value, end="\t")
        #             print("execute Found in row:", row[0].row)
        #             data.append(row)
        return data

    @pytest.mark.parametrize("data", read_excel_data("Testdata"))
    def test_login(self,Test_box, data):
        try:

            self. driver.find_element(By.XPATH, "//input[@name='uid']").send_keys(data["User ID"])

            self.driver.find_element(By.XPATH, "//input[@name='password']").send_keys(data["Password"])

            self.driver.find_element(By.XPATH, "//input[@name='btnLogin']").click()

            assert self.driver.current_url == "https://demo.guru99.com/V1/html/Managerhomepage.php"

        except Exception as error:
            print(error)
            pytest.fail(str(error))
            assert False
            # timestamp = str(int(time.time()))
            # directory = "screenshots"
            # if not os.path.exists(directory):
            #     os.makedirs(directory)
            # filename = f"{directory}/screenshot-{timestamp}.png"
            # self.driver.save_screenshot(filename)
            # print(f"Screenshot saved as {filename}")

