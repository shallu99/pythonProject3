import os
from datetime import time

import pytest
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By


class Test_do():
    driver = None
    @pytest.fixture()
    def Test_box(self):

        self.driver = webdriver.Chrome("C:\\webdriver\\chromedriver_win32\\chromedriver.exe")
        self.driver.get("https://demo.guru99.com/V1/index.php")
        self.driver.implicitly_wait(10)

        yield

        self.driver.quit()
        print("test completed")

    def read_excel_data(*sheet_names):
        wb = openpyxl.load_workbook("C:\\Users\\Shreyank M\\OneDrive\\Desktop\\guru99 data.xlsx")
        data = {}
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            rows = sheet.max_row
            cols = sheet.max_column
            sheet_data = []
            for r in range(2, rows + 1):
                row_data = {}
                for c in range(1, cols + 1):
                    row_data[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
                sheet_data.append(row_data)
            # data[sheet_name] = sheet_data
            print(data)
        return  sheet_data

    @pytest.mark.parametrize("login_data", read_excel_data("Testdata", "Testcase"))
    def test_login(self,Test_box, login_data):
        user_id = login_data[0]["User ID"]
        password = login_data[0]["Password"]


        try:
            # Login with the user ID and password
            self.driver.find_element(By.XPATH, "//input[@name='uid']").send_keys(user_id)
            self.driver.find_element(By.XPATH, "//input[@name='password']").send_keys(password)
            self.driver.find_element(By.XPATH, "//input[@name='btnLogin']").click()


        except Exception as error:
            print(error)
            timestamp = str(int(time.time()))
            directory = "screenshots"
            if not os.path.exists(directory):
                os.makedirs(directory)
            filename = f"{directory}/screenshot-{timestamp}.png"
            self.driver.save_screenshot(filename)
            print(f"Screenshot saved as {filename}")
