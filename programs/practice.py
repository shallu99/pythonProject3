import pytest
import openpyxl

from programs.basic_test import Test_do


class Test_do():
    driver = None
    fp = "C:\\Users\\Shreyank M\\OneDrive\\Desktop\\guru99 data.xlsx"

    @pytest.fixture()
    def Test_box(self, browser):
        self.driver = browser
        yield
        try:
            self.driver.quit()
            print("Test completed")
        except:
            print("Test completed")

    def new_read_excel_data(self, filepath, sheetname):
        wb = openpyxl.load_workbook(filepath)
        sheet = wb[sheetname]
        rows = sheet.max_row
        cols = sheet.max_column
        data = []
        for number_of_rowsInSheet in range(2, rows + 1):
            row_data = {}
            for number_of_columnsInSheet in range(1, cols + 1):
                value = sheet.cell(row=number_of_rowsInSheet, column=number_of_columnsInSheet).value
                if value is not None:
                    row_data[sheet.cell(row=1, column=number_of_columnsInSheet).value] = value
            if row_data.get("Execute", "") == "Yes":
                data.append(row_data)
        return data

    @staticmethod
    def extract(self):
        combined_dict = {}

        TestCase = self.new_read_excel_data(self.fp, 'TestCases')
        Env = self.new_read_excel_data(self.fp, 'Environment')
        Devices = self.new_read_excel_data(self.fp, 'Campaign(s)')

        output = []
        for item1 in TestCase:
            for item2 in Env:
                for item3 in Devices:
                    output.append([item1, item2, item3])

        combined_list = []
        for combination in output:

            combined_dict = {}
            for d in combination:
                combined_dict.update(d)
            combined_list.append(combined_dict)

        my_list = list(combined_dict.items())

        return combined_list

    @pytest.mark.parametrize("data", extract.__func__(Test_do))
    @pytest.mark.html
    @pytest.mark.hookwrapper
    def test_Cases(self, Test_box, data):
        try:
            self.driver.get(data["Login URL"])

        except AssertionError as error:
            print(error)
            assert False
        except Exception as error:
            print(error)
            assert False
