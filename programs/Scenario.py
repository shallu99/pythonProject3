import os
import time

import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl

from programs.basic_test import Test_do


class Test_do1():
    driver = None


    @pytest.fixture()
    def Test_box(self):

        self.driver = webdriver.Chrome("C:\\webdriver\\chromedriver_win32\\chromedriver.exe")
        self.driver.get("https://demo.guru99.com/V1/index.php")
        self.driver.implicitly_wait(10)

        yield

        self.driver.quit()
        print("test completed")

    def read_excel_data(filepath, Testdata):
        wb = openpyxl.load_workbook("C:\\Users\\Shreyank M\\OneDrive\\Desktop\\guru99 data.xlsx")
        sheet = wb[Testdata]
        rows = sheet.max_row
        cols = sheet.max_column
        data = []
        for r in range(2, rows + 1):
            row_data = {}
            for c in range(1, cols + 1):
                row_data[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
            data.append(row_data)
        return data

    @pytest.mark.parametrize("sheetname", ["Testdata", "Testcase"])
    def test_login(self, Test_box, sheetname):
        data = Test_do.read_excel_data("C:\\Users\\Shreyank M\\OneDrive\\Desktop\\guru99 data.xlsx", sheetname)

        try:
            for row in data:
                self.driver.find_element(By.XPATH, "//input[@name='uid']").send_keys(row["User ID"])
                self.driver.find_element(By.XPATH, "//input[@name='password']").send_keys(row["Password"])
                self.driver.find_element(By.XPATH, "//input[@name='btnLogin']").click()
                assert self.driver.current_url == "https://demo.guru99.com/V1/html/Managerhomepage.php"

        except Exception as error:
            print(error)
            timestamp = str(int(time.time()))
            directory = "screenshots"
            if not os.path.exists(directory):
                os.makedirs(directory)
            filename = f"{directory}/screenshot-{timestamp}.png"
            self.driver.save_screenshot(filename)
            print(f"Screenshot saved as {filename}")

            # data1 = Test_do.read_excel_data("C:\\Users\\Shreyank M\\OneDrive\\Desktop\\guru99 data.xlsx", sheetname)
            # try:
            #     for row in data1:
            #          assert self.driver.current_url == "https://demo.guru99.com/V1/index.php"
            #
            # except Exception as error:
            #     print(error)