import os

import openpyxl
import itertools
import configparser
class exceltables():
    def __init__(self):
        self.numOfColumns = 0
        # fetch current directory
        current_dir = os.getcwd()
        # Remove the "utilis" directory from the path
        parent_dir = os.path.dirname(os.path.dirname(current_dir))
        config = configparser.ConfigParser()
        # read properties from config file
        ConfigProperties_Path = parent_dir + "\\pytest\\utilis\\config.properties"
        config.read(ConfigProperties_Path)
        self.SHEET_TC = config.get('Test_ExcelData', 'SHEET_TC')
        self.SHEET_CAMPAIGNS_TOTEST = config.get('Test_ExcelData', 'SHEET_CAMPAIGNS_TOTEST')
        self.SHEET_ENVIRONMENTS_USERINPUT_LOGIN = config.get('Test_ExcelData', 'SHEET_ENVIRONMENTS_USERINPUT_LOGIN')
        # get file path and file name from properties
        Test_data_path = config.get('Test_ExcelData', 'Test_data_path')
        file_name = config.get('Test_ExcelData', 'file_name')
        test_data_file_path = parent_dir + "\\" + Test_data_path + "\\" + file_name
        # Create an object of the workbook class to read the excel file
        self.wb = openpyxl.load_workbook(test_data_file_path)

    def readExcelSheetOfTestCases(self,Classifier):
        self.tables = []
        # Get the sheet object
        sheet =self.wb[self.SHEET_TC]
        # Find the row index of the first occurrence of the search term
        firstRowIndex = 1
        # Calculate the number of columns
        self.numOfColumns = sheet.max_column
        # Store the table data in a dictionary
        table = []
        header = [cell.value for cell in sheet[firstRowIndex]]
        for i in range(firstRowIndex + 1,sheet.max_row + 1):
            row = {header[j]: sheet.cell(row=i, column=j + 1).value for j in range(self.numOfColumns)}
            if row['Classifier'] == Classifier :  # Check if the Execute value is 'No'
                # Remove key-value pairs where value is None or null
                row = {k: v for k, v in row.items() if v is not None}
                table.append(row)
        # Add the table data to the tables dictionary
        self.tables.append(table)
        print(self.tables)
        return self.tables

    def readExcelSheetOfCampaign(self,CAMPAIGNS):
        self.tables = []
        # Get the sheet object
        sheet = self.wb[self.SHEET_CAMPAIGNS_TOTEST]
        # Find the row index of the first occurrence of the search term
        firstRowIndex = 1
        # Calculate the number of columns
        self.numOfColumns = sheet.max_column
        # Store the table data in a dictionary
        table = []
        header = [cell.value for cell in sheet[firstRowIndex]]
        for i in range(firstRowIndex + 1, sheet.max_row + 1):
            row = {header[j]: sheet.cell(row=i, column=j + 1).value for j in range(self.numOfColumns)}
            if row['EXECUTE'] == 'Yes' and row['CAMPAIGNS'] == CAMPAIGNS:  # Check if the Execute value is 'No'
                # Remove key-value pairs where value is None or null
                row = {k: v for k, v in row.items() if v is not None}
                table.append(row)
        # Add the table data to the tables dictionary
        self.tables.append(table)
        print(self.tables)
        return self.tables

    def readExcelSheetOfEnvironment(self):
        self.tables = []
        # Get the sheet object
        sheet =self.wb[self.SHEET_ENVIRONMENTS_USERINPUT_LOGIN]
        # Find the row index of the first occurrence of the search term
        firstRowIndex = 1
        # Calculate the number of columns
        self.numOfColumns = sheet.max_column
        # Store the table data in a dictionary
        table = []
        header = [cell.value for cell in sheet[firstRowIndex]]
        for i in range(firstRowIndex + 1, sheet.max_row + 1):
            row = {header[j]: sheet.cell(row=i, column=j + 1).value for j in range(self.numOfColumns)}
            if row['Execute'] == 'Yes':  # Check if the Execute value is 'No'
                # Remove key-value pairs where value is None or null
                row = {k: v for k, v in row.items() if v is not None}
                table.append(row)
        # Add the table data to the tables dictionary
        self.tables.append(table)
        return self.tables

    def readExcelSheetcombinations(self,Classifier):
        CAMPAIGNS = Classifier
        # Get the set of all keys in all dictionaries
        read = exceltables()
        self.table1 = read.readExcelSheetOfEnvironment()
        self.table2= read.readExcelSheetOfCampaign(CAMPAIGNS)
        self.table3 = read.readExcelSheetOfTestCases(Classifier)
        self.data = []
        for test_case in self.table1:
            self.data.append(test_case)
        for test_case in self.table2:
            self.data.append(test_case)
        for test_case in self.table3:
            self.data.append(test_case)
        all_keys = set()
        for table in self.data:
            for row in table:
                all_keys |= set(row.keys())
        # Create a new dictionary for each combination with default values of None
        combinations = []
        for table in itertools.product(*self.data):
            new_dict = {key: None for key in all_keys}
            for row in table:
                new_dict.update(row)
            combinations.append(new_dict)
        print("readExcelSheetcombinations ",combinations)
        return combinations
if __name__ == '__main__':
    objExceltable = exceltables()
    objExceltable.readExcelSheetcombinations('T001')